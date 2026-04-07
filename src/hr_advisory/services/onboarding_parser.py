"""Onboarding template Excel parser.

Parses the LIA Onboarding Templates .xlsx format (12 sheets) and returns
structured data for creating onboarding modules and steps.  Handles missing
optional sheets gracefully, validates file size and format, sanitises all
string content, and ignores formula cells (data_only mode).
"""

from __future__ import annotations

import io
import logging
from typing import Any

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
_MAX_STRING_LENGTH = 5_000
_XLSX_MAGIC = b"PK"  # ZIP/OOXML files start with PK header

# Sheet names as they appear in the template workbook.  The "Instructions"
# sheet is intentionally skipped -- it contains guidance, not data.
_SHEET_COMPANY_PROFILE = "1. Company Profile"
_SHEET_ORG_STRUCTURE = "2. Org Structure"
_SHEET_MODULES = "3. Onboarding Modules"
_SHEET_STEP_CONTENT = "4. Step Content"
_SHEET_ROLE_CONFIG = "5. Role Configuration"
_SHEET_IT_PROVISIONING = "6. IT Provisioning"
_SHEET_POLICIES = "7. Policies & Compliance"
_SHEET_BENEFITS = "8. Benefits Overview"
_SHEET_PROBATION = "9. Probation & Goals"
_SHEET_COMMS = "10. Comms & Channels"
_SHEET_CONTACTS = "11. Key Contacts"
_SHEET_PREBOARDING = "12. Pre-boarding Checklist"

_REQUIRED_SHEETS = {_SHEET_MODULES, _SHEET_STEP_CONTENT}

# Company Profile key-value field mapping.
# Maps the "Field" column label to the output dict key.
_COMPANY_PROFILE_FIELDS: dict[str, str] = {
    "Company Name": "company_name",
    "Industry / Sector": "industry",
    "Founded": "founded",
    "Headquarters": "headquarters",
    "Number of Employees": "employee_count",
    "Office Locations": "office_locations",
    "Mission Statement": "mission",
    "Vision Statement": "vision",
    "Company History (Brief)": "company_history",
    "Core Value 1": "core_value_1",
    "Core Value 2": "core_value_2",
    "Core Value 3": "core_value_3",
    "Core Value 4": "core_value_4",
    "Core Value 5": "core_value_5",
    "CEO / Founder Name": "ceo_name",
    "Website URL": "website",
    "Company LinkedIn": "linkedin",
    "Brand Tone of Voice": "tone_of_voice",
}

# Probation & Goals key-value field mapping.
_PROBATION_FIELDS: dict[str, str] = {
    "Standard Probation Length": "probation_length",
    "Probation Review Schedule": "review_schedule",
    "Review Format": "review_format",
    "Performance Rating Scale": "rating_scale",
    "Probation Extension Policy": "extension_policy",
    "Successful Completion Criteria": "completion_criteria",
    "Who Conducts Reviews?": "reviewer",
    "Documentation Required?": "documentation_required",
    "30-Day Goal Template": "goal_30_day",
    "60-Day Goal Template": "goal_60_day",
    "90-Day Goal Template": "goal_90_day",
    "Goal-Setting Framework": "goal_framework",
    "Feedback Tool / Platform": "feedback_tool",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _sanitise(value: Any) -> Any:
    """Sanitise a cell value.

    - Strings are stripped of leading/trailing whitespace and truncated to
      ``_MAX_STRING_LENGTH``.
    - Numbers and booleans pass through unchanged.
    - ``None`` is returned as-is.
    """
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if len(cleaned) > _MAX_STRING_LENGTH:
            cleaned = cleaned[:_MAX_STRING_LENGTH]
        return cleaned if cleaned else None
    return value


def _row_is_empty(row: tuple) -> bool:
    """Return True if every cell in the row is None or blank."""
    return all(
        v is None or (isinstance(v, str) and not v.strip())
        for v in row
    )


def _split_comma_list(value: Any) -> list[str]:
    """Split a comma-separated string into a list of stripped, non-empty items."""
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def _parse_key_value_sheet(
    ws: Any,
    field_map: dict[str, str],
    header_row: int = 5,
) -> dict[str, Any]:
    """Parse a key-value sheet where column A is the field label and column B
    is the value.

    ``header_row`` is the 1-based row containing "Field" / "Your Organisation's
    Details" headers.  Data rows start at ``header_row + 1``.
    """
    result: dict[str, Any] = {}
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        max_col=2,
        values_only=True,
    ):
        if _row_is_empty(row):
            continue
        label = _sanitise(row[0])
        value = _sanitise(row[1]) if len(row) > 1 else None
        if label and label in field_map:
            result[field_map[label]] = value
    return result


def _parse_table_sheet(
    ws: Any,
    column_keys: list[str],
    header_row: int = 4,
    comma_split_columns: set[str] | None = None,
) -> list[dict[str, Any]]:
    """Parse a tabular sheet into a list of dicts.

    ``header_row`` is the 1-based row containing column headers.  Data rows
    start at ``header_row + 1``.  ``column_keys`` defines the output dict
    keys in column order.

    If ``comma_split_columns`` is provided, those keys will have their
    values split on commas into a list.
    """
    if comma_split_columns is None:
        comma_split_columns = set()

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        max_col=len(column_keys),
        values_only=True,
    ):
        if _row_is_empty(row):
            continue
        record: dict[str, Any] = {}
        for idx, key in enumerate(column_keys):
            raw = row[idx] if idx < len(row) else None
            if key in comma_split_columns:
                record[key] = _split_comma_list(raw)
            else:
                record[key] = _sanitise(raw)
        rows.append(record)
    return rows


# ---------------------------------------------------------------------------
# Individual sheet parsers
# ---------------------------------------------------------------------------


def _parse_company_profile(wb: Any, warnings: list[str]) -> dict[str, Any]:
    """Parse Sheet 1 — Company Profile (key-value pairs)."""
    if _SHEET_COMPANY_PROFILE not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_COMPANY_PROFILE}' not found, skipping.")
        return {}

    ws = wb[_SHEET_COMPANY_PROFILE]
    result = _parse_key_value_sheet(ws, _COMPANY_PROFILE_FIELDS, header_row=5)

    # Collect core values into a list
    core_values: list[str] = []
    for i in range(1, 6):
        key = f"core_value_{i}"
        val = result.pop(key, None)
        if val:
            core_values.append(val)
    result["core_values"] = core_values

    return result


def _parse_org_structure(wb: Any, warnings: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 2 — Org Structure (table)."""
    if _SHEET_ORG_STRUCTURE not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_ORG_STRUCTURE}' not found, skipping.")
        return []

    ws = wb[_SHEET_ORG_STRUCTURE]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "department_name",
            "department_head",
            "head_title",
            "description",
            "team_size",
            "contact_email",
        ],
        header_row=4,
    )


def _parse_modules(wb: Any, errors: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 3 — Onboarding Modules (table, REQUIRED)."""
    if _SHEET_MODULES not in wb.sheetnames:
        errors.append(f"Required sheet '{_SHEET_MODULES}' not found.")
        return []

    ws = wb[_SHEET_MODULES]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "module_number",
            "module_name",
            "description",
            "steps",
            "duration",
            "phase",
            "required",
            "role_specific",
        ],
        header_row=5,
        comma_split_columns={"steps"},
    )


def _parse_step_content(wb: Any, errors: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 4 — Step Content (table, REQUIRED)."""
    if _SHEET_STEP_CONTENT not in wb.sheetnames:
        errors.append(f"Required sheet '{_SHEET_STEP_CONTENT}' not found.")
        return []

    ws = wb[_SHEET_STEP_CONTENT]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "module_name",
            "step_name",
            "step_number",
            "heading",
            "body_content",
            "checklist_items",
            "media",
        ],
        header_row=4,
        comma_split_columns={"checklist_items"},
    )


def _parse_role_configs(wb: Any, warnings: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 5 — Role Configuration (table)."""
    if _SHEET_ROLE_CONFIG not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_ROLE_CONFIG}' not found, skipping.")
        return []

    ws = wb[_SHEET_ROLE_CONFIG]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "role",
            "department",
            "hiring_manager",
            "additional_modules",
            "tools",
            "buddy_assignment",
            "probation_length",
            "goals_summary",
        ],
        header_row=4,
        comma_split_columns={"additional_modules", "tools"},
    )


def _parse_it_provisioning(wb: Any, warnings: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 6 — IT Provisioning (table)."""
    if _SHEET_IT_PROVISIONING not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_IT_PROVISIONING}' not found, skipping.")
        return []

    ws = wb[_SHEET_IT_PROVISIONING]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "tool",
            "category",
            "applicable_roles",
            "provisioning_owner",
            "sla",
            "access_level",
            "setup_url",
            "notes",
        ],
        header_row=4,
        comma_split_columns={"applicable_roles"},
    )


def _parse_policies(wb: Any, warnings: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 7 — Policies & Compliance (table)."""
    if _SHEET_POLICIES not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_POLICIES}' not found, skipping.")
        return []

    ws = wb[_SHEET_POLICIES]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "policy_name",
            "category",
            "description",
            "acknowledgement_required",
            "document_url",
            "review_deadline",
            "applicable_to",
            "renewal_frequency",
        ],
        header_row=4,
    )


def _parse_benefits(wb: Any, warnings: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 8 — Benefits Overview (table)."""
    if _SHEET_BENEFITS not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_BENEFITS}' not found, skipping.")
        return []

    ws = wb[_SHEET_BENEFITS]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "category",
            "benefit_name",
            "description",
            "eligibility",
            "enrollment_deadline",
            "provider",
            "contact_url",
        ],
        header_row=4,
    )


def _parse_probation_config(wb: Any, warnings: list[str]) -> dict[str, Any]:
    """Parse Sheet 9 — Probation & Goals (key-value pairs)."""
    if _SHEET_PROBATION not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_PROBATION}' not found, skipping.")
        return {}

    ws = wb[_SHEET_PROBATION]
    # Header row is row 4, but row 5 is empty; data starts at row 6
    return _parse_key_value_sheet(ws, _PROBATION_FIELDS, header_row=4)


def _parse_communications(wb: Any, warnings: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 10 — Comms & Channels (table)."""
    if _SHEET_COMMS not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_COMMS}' not found, skipping.")
        return []

    ws = wb[_SHEET_COMMS]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "channel",
            "type",
            "purpose",
            "frequency",
            "who_involved",
            "platform",
            "access_instructions",
        ],
        header_row=4,
    )


def _parse_contacts(wb: Any, warnings: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 11 — Key Contacts (table)."""
    if _SHEET_CONTACTS not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_CONTACTS}' not found, skipping.")
        return []

    ws = wb[_SHEET_CONTACTS]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "name",
            "role",
            "department",
            "email",
            "available_as_buddy",
            "buddy_capacity",
            "notes",
        ],
        header_row=4,
    )


def _parse_preboarding(wb: Any, warnings: list[str]) -> list[dict[str, Any]]:
    """Parse Sheet 12 — Pre-boarding Checklist (table)."""
    if _SHEET_PREBOARDING not in wb.sheetnames:
        warnings.append(f"Optional sheet '{_SHEET_PREBOARDING}' not found, skipping.")
        return []

    ws = wb[_SHEET_PREBOARDING]
    return _parse_table_sheet(
        ws,
        column_keys=[
            "task",
            "owner",
            "trigger",
            "deadline",
            "auto_assignable",
            "dependency",
            "notes",
        ],
        header_row=4,
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse_onboarding_template(file_bytes: bytes) -> dict[str, Any]:
    """Parse a LIA onboarding template .xlsx file.

    Accepts raw file bytes, validates the file, parses all 12 data sheets,
    and returns structured data suitable for creating onboarding modules
    and steps in the system.

    Args:
        file_bytes: The raw bytes of the .xlsx file.

    Returns:
        A dict containing::

            {
                "company_profile": {...},
                "departments": [...],
                "modules": [...],
                "steps": [...],
                "role_configs": [...],
                "it_tasks": [...],
                "policies": [...],
                "benefits": [...],
                "probation_config": {...},
                "communications": [...],
                "contacts": [...],
                "preboarding_tasks": [...],
                "warnings": [...],
                "errors": [...],
            }

        ``warnings`` contains non-fatal issues (missing optional sheets,
        skipped rows).  ``errors`` contains fatal issues (missing required
        sheets, invalid file format).  When ``errors`` is non-empty, the
        data sections may be empty or incomplete.

    Raises:
        ValueError: If the file exceeds the size limit or is not a valid
            .xlsx file.
    """
    warnings: list[str] = []
    errors: list[str] = []

    # -- Validate file size --
    if len(file_bytes) > _MAX_FILE_SIZE:
        raise ValueError(
            f"File size ({len(file_bytes):,} bytes) exceeds the "
            f"{_MAX_FILE_SIZE // (1024 * 1024)} MB limit."
        )

    # -- Validate MIME / magic bytes --
    if not file_bytes[:2].startswith(_XLSX_MAGIC):
        raise ValueError(
            "File does not appear to be a valid .xlsx file "
            "(expected ZIP/OOXML format)."
        )

    # -- Load workbook --
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for parsing .xlsx files. "
            "Install with: pip install openpyxl>=3.1.0"
        )

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise ValueError(f"Failed to open .xlsx file: {exc}") from exc

    try:
        # -- Check required sheets --
        available = set(wb.sheetnames)
        for required_sheet in _REQUIRED_SHEETS:
            if required_sheet not in available:
                errors.append(f"Required sheet '{required_sheet}' not found.")

        if errors:
            # Still attempt to parse whatever is available
            logger.warning(
                "Required sheets missing from onboarding template: %s",
                errors,
            )

        # -- Parse each sheet --
        company_profile = _parse_company_profile(wb, warnings)
        departments = _parse_org_structure(wb, warnings)
        modules = _parse_modules(wb, errors)
        steps = _parse_step_content(wb, errors)
        role_configs = _parse_role_configs(wb, warnings)
        it_tasks = _parse_it_provisioning(wb, warnings)
        policies = _parse_policies(wb, warnings)
        benefits = _parse_benefits(wb, warnings)
        probation_config = _parse_probation_config(wb, warnings)
        communications = _parse_communications(wb, warnings)
        contacts = _parse_contacts(wb, warnings)
        preboarding_tasks = _parse_preboarding(wb, warnings)

        # -- Summary logging --
        logger.info(
            "Parsed onboarding template: %d modules, %d steps, "
            "%d departments, %d role configs, %d warnings, %d errors",
            len(modules),
            len(steps),
            len(departments),
            len(role_configs),
            len(warnings),
            len(errors),
        )

        return {
            "company_profile": company_profile,
            "departments": departments,
            "modules": modules,
            "steps": steps,
            "role_configs": role_configs,
            "it_tasks": it_tasks,
            "policies": policies,
            "benefits": benefits,
            "probation_config": probation_config,
            "communications": communications,
            "contacts": contacts,
            "preboarding_tasks": preboarding_tasks,
            "warnings": warnings,
            "errors": errors,
        }
    finally:
        wb.close()
