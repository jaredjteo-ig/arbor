"""Regression + unit tests for the onboarding template Excel import endpoint (T195).

Covers ``POST /onboarding/templates/import``:

1. Minimum-viable import — only sheets 3 (Modules) and 4 (Step Content) are
   present. The endpoint must still create a template plus its modules and
   steps.
2. Partial-import warnings — a step that references an unknown module is
   recovered (assigned to the first module) and reported in ``warnings``.
3. Tenant isolation — the ``company_id`` written to all created records must
   come from the authenticated user, NEVER from any value embedded in the
   uploaded workbook.

The tests build a real .xlsx workbook in memory using openpyxl (already a
project dependency) and exercise the endpoint via FastAPI's TestClient with
``dataflow_crud`` mocked. No database or HTTP I/O is performed.
"""

from __future__ import annotations

import io
from typing import Any
from unittest.mock import MagicMock, patch

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

import openpyxl

from hr_advisory.api.middleware.auth_middleware import get_current_user
from hr_advisory.api.middleware.rate_limit import reset_rate_limit_state
from hr_advisory.api.routers.onboarding import router


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _fake_owner() -> dict:
    return {
        "sub": "10",
        "email": "owner@example.com",
        "role": "owner",
        "company_id": 42,
    }


def _make_app() -> FastAPI:
    app = FastAPI()
    app.include_router(router, prefix="/onboarding")
    return app


@pytest.fixture()
def owner_client():
    """Test client authenticated as company owner (company_id=42)."""
    reset_rate_limit_state()
    app = _make_app()
    app.dependency_overrides[get_current_user] = _fake_owner
    yield TestClient(app)
    app.dependency_overrides.clear()
    reset_rate_limit_state()


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------

# Header row is row 5 for "3. Onboarding Modules"; data begins at row 6.
_MODULE_HEADER_ROW = 5
# Header row is row 4 for "4. Step Content"; data begins at row 5.
_STEP_HEADER_ROW = 4


def _add_modules_sheet(wb: openpyxl.Workbook, modules: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("3. Onboarding Modules")
    # Filler rows above the header — the parser skips them.
    for _ in range(_MODULE_HEADER_ROW - 1):
        ws.append([None])
    ws.append([
        "Module #", "Module Name", "Description", "Steps",
        "Duration (min)", "Phase", "Required", "Role-specific",
    ])
    for idx, mod in enumerate(modules, start=1):
        ws.append([
            idx,
            mod["name"],
            mod.get("description", ""),
            mod.get("steps_csv", ""),
            mod.get("duration", 30),
            mod.get("phase", "orientation"),
            "Yes" if mod.get("required", True) else "No",
            "Yes" if mod.get("role_specific", False) else "No",
        ])


def _add_step_content_sheet(wb: openpyxl.Workbook, steps: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("4. Step Content")
    for _ in range(_STEP_HEADER_ROW - 1):
        ws.append([None])
    ws.append([
        "Module Name", "Step Name", "Step #", "Heading",
        "Body Content", "Checklist Items", "Media",
    ])
    for step in steps:
        ws.append([
            step.get("module_name", ""),
            step.get("step_name", ""),
            step.get("step_number", 1),
            step.get("heading", step.get("step_name", "")),
            step.get("body_content", ""),
            step.get("checklist_items", ""),
            step.get("media", ""),
        ])


def _build_workbook(
    modules: list[dict[str, Any]],
    steps: list[dict[str, Any]],
    *,
    company_name: str | None = None,
) -> bytes:
    """Build an in-memory .xlsx file mimicking the LIA template shape."""
    wb = openpyxl.Workbook()
    # Remove the default sheet — parser does not need it.
    default_name = wb.sheetnames[0]
    del wb[default_name]

    if company_name is not None:
        ws = wb.create_sheet("1. Company Profile")
        # Header at row 5; data rows after.
        for _ in range(4):
            ws.append([None])
        ws.append(["Field", "Your Organisation's Details"])
        ws.append(["Company Name", company_name])

    _add_modules_sheet(wb, modules)
    _add_step_content_sheet(wb, steps)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# dataflow_crud mock helper
# ---------------------------------------------------------------------------


class _FakeCRUD:
    """In-memory replacement for ``dataflow_crud`` operations.

    Mirrors the public surface used by the import endpoint: ``create``,
    ``read``, ``list_records``, ``update``. Records are auto-assigned
    incrementing ids per model.
    """

    def __init__(self) -> None:
        self.records: dict[str, list[dict[str, Any]]] = {}
        self._next_id = 1

    def create(self, model_name: str, data: dict[str, Any]) -> dict[str, Any]:
        rec = dict(data)
        rec["id"] = self._next_id
        self._next_id += 1
        self.records.setdefault(model_name, []).append(rec)
        return rec

    def read(self, model_name: str, record_id: int) -> dict[str, Any] | None:
        for rec in self.records.get(model_name, []):
            if rec.get("id") == record_id:
                return dict(rec)
        return None

    def list_records(
        self,
        model_name: str,
        filter_dict: dict[str, Any] | None = None,
        limit: int | None = None,
    ) -> list[dict[str, Any]]:
        rows = self.records.get(model_name, [])
        if not filter_dict:
            return [dict(r) for r in rows]
        out: list[dict[str, Any]] = []
        for r in rows:
            if all(r.get(k) == v for k, v in filter_dict.items()):
                out.append(dict(r))
        return out

    def update(self, model_name: str, record_id: int, data: dict[str, Any]) -> dict[str, Any]:
        for r in self.records.get(model_name, []):
            if r.get("id") == record_id:
                r.update(data)
                return dict(r)
        raise KeyError(f"{model_name}#{record_id} not found")


@pytest.fixture()
def fake_crud():
    crud = _FakeCRUD()
    with patch(
        "hr_advisory.api.routers.onboarding.dataflow_crud",
        new=MagicMock(
            create=crud.create,
            read=crud.read,
            list_records=crud.list_records,
            update=crud.update,
        ),
    ):
        yield crud


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestOnboardingTemplateImport:
    """POST /onboarding/templates/import — Excel import end-to-end."""

    def test_minimum_viable_import_creates_template_modules_and_steps(
        self, owner_client, fake_crud
    ):
        """Sheets 3 + 4 alone yield a usable template with modules and steps."""
        modules = [
            {"name": "Welcome to Acme", "phase": "orientation", "duration": 45},
            {"name": "Compliance Basics", "phase": "compliance", "duration": 60},
        ]
        steps = [
            {
                "module_name": "Welcome to Acme",
                "step_name": "Company Mission",
                "step_number": 1,
                "heading": "Our Mission",
                "body_content": "We make great software.",
            },
            {
                "module_name": "Compliance Basics",
                "step_name": "Code of Conduct",
                "step_number": 1,
                "heading": "Code of Conduct",
                "body_content": "Read and follow these rules.",
                "checklist_items": "Read; Sign; Confirm",
            },
        ]
        xlsx_bytes = _build_workbook(modules, steps)

        resp = owner_client.post(
            "/onboarding/templates/import",
            files={
                "file": (
                    "acme.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["template_id"] is not None
        assert body["modules_created"] == 2
        # Two parsed steps from sheet 4. No optional sheets were provided so
        # there should be no extra synthesised steps.
        assert body["steps_created"] == 2
        assert body["preboarding_tasks_created"] == 0

        # Storage assertions
        assert len(fake_crud.records["OnboardingTemplate"]) == 1
        assert len(fake_crud.records["OnboardingModule"]) == 2
        assert len(fake_crud.records["OnboardingStep"]) == 2

        # Steps are linked to the right module
        modules_by_name = {m["name"]: m for m in fake_crud.records["OnboardingModule"]}
        welcome_id = modules_by_name["Welcome to Acme"]["id"]
        compliance_id = modules_by_name["Compliance Basics"]["id"]
        step_module_ids = {s["module_id"] for s in fake_crud.records["OnboardingStep"]}
        assert step_module_ids == {welcome_id, compliance_id}

    def test_step_with_unknown_module_recovers_with_warning(
        self, owner_client, fake_crud
    ):
        """A step referencing an unknown module is reassigned to the first
        module and a warning is reported. The import does NOT abort."""
        modules = [
            {"name": "Welcome", "phase": "orientation", "duration": 30},
        ]
        steps = [
            {
                "module_name": "Welcome",
                "step_name": "Intro",
                "heading": "Intro",
                "body_content": "Hello",
            },
            {
                "module_name": "Does Not Exist",
                "step_name": "Orphan",
                "heading": "Orphan",
                "body_content": "This step has a bad parent.",
            },
        ]
        xlsx_bytes = _build_workbook(modules, steps)

        resp = owner_client.post(
            "/onboarding/templates/import",
            files={
                "file": (
                    "acme.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["modules_created"] == 1
        # Both steps still get created — the orphan attaches to the first
        # module rather than failing the whole import.
        assert body["steps_created"] == 2

        # All steps are attached to the only existing module.
        only_mod_id = fake_crud.records["OnboardingModule"][0]["id"]
        for step in fake_crud.records["OnboardingStep"]:
            assert step["module_id"] == only_mod_id

    def test_tenant_isolation_uses_auth_company_id_not_workbook_company(
        self, owner_client, fake_crud
    ):
        """The ``company_id`` on every created record MUST come from the
        authenticated user (``42``). A different company name embedded in the
        workbook must NOT bleed into the records."""
        modules = [
            {"name": "Welcome", "phase": "orientation", "duration": 30},
        ]
        steps = [
            {
                "module_name": "Welcome",
                "step_name": "Intro",
                "heading": "Intro",
                "body_content": "Hi",
            },
        ]
        # The workbook claims to belong to a *different* company. The endpoint
        # must ignore this and use the authenticated user's company_id.
        xlsx_bytes = _build_workbook(
            modules, steps, company_name="Other Tenant Inc"
        )

        resp = owner_client.post(
            "/onboarding/templates/import",
            files={
                "file": (
                    "import.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 200, resp.text

        tmpl = fake_crud.records["OnboardingTemplate"][0]
        assert tmpl["company_id"] == 42, (
            "Template company_id must be derived from auth context (42), "
            f"not from the workbook contents — got {tmpl['company_id']}"
        )
        for mod in fake_crud.records["OnboardingModule"]:
            assert mod["company_id"] == 42

    def test_rejects_non_xlsx_file_extension(self, owner_client, fake_crud):
        """Non-.xlsx uploads are rejected with HTTP 400 and no records created."""
        resp = owner_client.post(
            "/onboarding/templates/import",
            files={"file": ("bad.csv", b"name,phase\nWelcome,orientation\n", "text/csv")},
        )
        assert resp.status_code == 400
        assert "xlsx" in resp.json()["detail"].lower()
        assert fake_crud.records == {}
